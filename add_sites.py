import os, re
import string
import sys

import pandas as pd
import requests
from datetime import datetime, date
from tqdm import tqdm
import openpyxl as pxl
from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copy, rmtree
from openpyxl.worksheet.datavalidation import DataValidation
import urllib.parse
from bs4 import BeautifulSoup
import time
import logging

from helpers import logg
from env import config


## Defining variables NOTE: user and password are kept in dotenv file. Only necessary if using NOTD api.
today = date.today().strftime("%d/%m/%Y")
working_dir = os.getcwd()
folder = 'New URLs between ' +  input('New URLs between> ') + '/'
metadata_folder = folder+'metadata/'

###SET UP - Creating folders
logger = logging.getLogger()
logger.setLevel(10)
logg.configure_handlers(logger, f'{metadata_folder}logs/')      #Configures logs and creates folder and metadata folder.
logger.addHandler(logg.create_StreamHandler(sys.stdout, 40, format='%(message)s'))

logger.debug('Creating Folders')
try:
    while len([x for x in os.listdir('Harvesting Summary') if x[-4:] == '.csv']) != 1:
        if len([x for x in os.listdir('Harvesting Summary') if x[-4:] == '.csv']) > 1:
            logger.error('Too many CSVs in Harvesting Summary folder')
        elif len([x for x in os.listdir('Harvesting Summary') if x[-4:] == '.csv']) < 1:
            logger.error('No CSVs in Harvesting Summary folder')
        input('\nPlease add Harvesting Summary csv to Harvesting Summary Directory and hit enter:>')

    source = f'Harvesting Summary/{[x for x in os.listdir("Harvesting Summary") if x[-4:] == ".csv"][0]}'

    os.replace(source, f'{metadata_folder}Harvesting Summary.csv')

except:
    logger.exception('Error, Aborting process and reverting to prior state.')
    [x.close() for x in logger.handlers]
    rmtree(folder)
    raise Exception('Process Aborted. Changes rolled back.')

def UKGWA_URL(url: str) -> str:
    """Creates a UKGWA URL from a standard URL
    (Includes Social Media URLs)"""
    social = {'twitter.com': 'twitter',
              'flickr.com': 'flickr',
              'youtube.com': 'video'}

    domain = re.search('.*:\/\/(?:www.)?([^\/]+)', url).group(1)
    logger.debug(f'url: {url}. Domain: {domain}')
    if domain in social:
        channel = url[:-1].split('/')[-1] + url[-1]     ### takes last character off before splitting in case it is a slash '/'. Appends after split
        return f'https://webarchive.nationalarchives.gov.uk/{social[domain]}/{channel}'.strip()
    else:
        return f'https://webarchive.nationalarchives.gov.uk/ukgwa/*/{url}'.strip()


def first_capture(archive_url: str) -> str:
    """Gets the year and month of the first capture of any URL
    Date is return in MONTH YEAR format e.g. 'July 2016'

    Function sleeps for 0.3 seconds to limit rate of requests."""
    time.sleep(0.3)
    logger.debug(f'Getting First Capture of {archive_url}')
    if '/*/' in archive_url:    ## CDX API Query
        url = archive_url.split('/*/')[1]
        #query = f'http://tnaqanotd.mirrorweb.com/published-cdx?url={urllib.parse.quote(url)}&fields=timestamp&limit=1' #out of date.
        query = f'https://webarchive.nationalarchives.gov.uk/ukgwa/cdx?url={urllib.parse.quote(url)}&fields=timestamp&limit=1&output=text'
        try:
            r = requests.get(query, allow_redirects=False)#, auth=(config.NOTDuser, config.NOTDpassword))
            logger.debug(f'{archive_url} response code: {r.status_code}')
            if r.status_code == 200:
                date = r.text[:6]
                date = datetime.strptime(date, '%Y%m')
                date_string = date.strftime('%B %Y')
                return date_string
        except:
            logger.warning(f'Request to {archive_url} failed.')
    else:
        logger.debug(f'Scraping Social Media page for earliest capture of {archive_url}')
        ## Scrapes social media page for first Date of capture.
        try:
            r = requests.get(archive_url+'?sort=date_oldest', timeout=30)
            soup = BeautifulSoup(r.text, 'html.parser')
            if '/twitter/' in archive_url:
                tag = soup.find('li')
                tag = tag.find(attrs={'class': 'mwtwtime'})
                tag = tag.find('span')
                date = tag.text.replace(' ', '').replace('\n', '')
            if '/video/' in archive_url:
                tag = soup.find('li')
                tag = tag.find(attrs={'class': 'video-date-s'})
                date = tag.text
            if '/flickr/' in archive_url:
                tag = soup.find(attrs={'class': 'col-sm-6 col-md-4'})
                tag = tag.find(attrs={'class': 'date'})
                date = tag.text.replace(' ', '').replace('\n', '')

            date = datetime.strptime(date, '%d/%m/%Y')
            date_string = date.strftime('%B %Y')
            return date_string
        except:
            logger.warning(f'Request to {archive_url} failed.')
    logger.info(f'{archive_url} not available.')
    return ''

####Load and clean Harvesting Summary
logger.debug('Loading Harvest Summary')
full_list = pd.read_excel(f'Full List.xlsx') ###IGNORE IF ALREADY IN LIST?
new_sites = pd.read_csv(f'{metadata_folder}Harvesting Summary.csv')
new_sites.loc[:, 'Additional Information'] = ''
new_sites.rename(columns={'textbox20': 'URL', 'textbox22': 'Site Name', 'textbox26': 'Archivist Notes', 'Dept_acronym': 'Department'}, inplace=True)
new_sites.loc[:, 'Archive URL'] = new_sites['URL'].apply(UKGWA_URL)
new_sites = new_sites[~new_sites['Archive URL'].isin(full_list['Archive URL'])]

#####BACK UPS
logger.debug(f'Creating backups in {metadata_folder}')
copy('not_active.csv', f'{metadata_folder}not_active prev.csv')
copy('Full List.xlsx', f'{metadata_folder}Full List prev.xlsx')

###Wrapped in try: except: so that an error rolls back.
try:
    ####CHECK URLs
    logger.debug('Creating Checklist of URLs')
    frames = [new_sites] + [pd.read_csv(file) for file in os.listdir() if 'not_active' in file]
    to_check = pd.concat(frames, ignore_index=True)
    to_check.drop_duplicates(subset='URL', inplace=True)
    logger.debug('Checking URLs')
    tqdm.pandas(desc='Checking URLs against Archive')       #Creates tqdm instance for progress bar
    to_check.loc[:, 'From'] = to_check['Archive URL'].progress_apply(first_capture)
    to_check.to_csv(f'{metadata_folder}All checked.csv', index=False)

    #####CREATE NOT ACTIVE CSV
    logger.debug('Saving non-active URLs to file not_active.csv')
    not_active = to_check[to_check['From'] == ''].drop(['From'], axis=1)
    not_active.to_csv('not_active.csv', index=False)

    ######CREATE ACTIVE URLs XLSX
    logger.debug('Configuring active_sites list')
    active_sites = to_check[to_check['From'] != ''].loc[:]
    active_sites.loc[:, 'To'] = 'Ongoing'
    active_sites = active_sites[['Archive URL', 'Site Name', 'From', 'To', 'Additional Information', 'Archivist Notes', 'Department']]  ####define columns
    for x in range(1, 7):
        active_sites[f'Category #{x}'] = ''
    active_sites.loc[:, 'Added to Full List'] = today
    wb = pxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(active_sites, index=False, header=True):
        ws.append(r)

    #### CREATE DATA VALIDATION FOR CATEGORIES
    logger.debug('Configuring Verification sheet.')
    cats = wb.create_sheet('cats')
    cats.sheet_state = 'hidden'
    categories = ['Business, industry, economics and finance', 'Central and regional government', 'Culture and leisure', 'Environment', 'Health, well-being and care', 'Home affairs, public order, justice and rights', 'Honours, awards and appointments', 'International affairs and defence', 'People, community and housing', 'Public inquiries, inquests and royal commissions', 'Transport, communication, science and technology', 'Work, education and skills']
    for i, x in enumerate(categories):
        cats.cell(row=i+1, column=1).value = x
    dv = DataValidation(type="list", formula1=f'cats!$A$1:$A${i+1}', allow_blank=True)

    ws.add_data_validation(dv)
    dv.add('H2:M1048576')
    wb.save(f'{folder}Verification.xlsx')

    logger.debug('User Verifying site names/categories.')
    input(f'''
Verify Site names and add categories to sites in '{folder}Verification.xlsx'
When finished, close and save spreadsheet and hit enter here:>''')
    input('\nRe-hit enter to confirm:>')

    ###Create cataloguing sheet
    logger.debug('Creating Cataloguing Sheet')
    verified = pd.read_excel(f'{folder}Verification.xlsx')
    copy(f'{folder}Verification.xlsx', f'{metadata_folder}Verified New sites.xlsx')
    cataloguing = verified.drop(columns=['Added to Full List', 'Archivist Notes','Category #1',
                                         'Category #2', 'Category #3', 'Category #4', 'Category #5',
                                         'Category #6'])
    cataloguing.to_excel(f'{folder}Verification.xlsx', index=False, )
    os.rename(f'{folder}Verification.xlsx', f'{folder}cataloguing.xlsx')
    to_full_list = verified[['Archive URL', 'Site Name', 'From', 'To',
                              'Department', 'Category #1', 'Category #2',
                              'Category #3', 'Category #4', 'Category #5',
                              'Category #6', 'Additional Information',
                              'Archivist Notes', 'Added to Full List']]

    ###ADD NEW LIST TO FULL LIST
    logger.debug('Updating Full list.xlsx')
    while True:
        try:
            frames = [full_list] + [to_full_list]
            full_list = pd.concat(frames, ignore_index=True)
            full_list.drop_duplicates(subset='Archive URL', keep='last', inplace=True) ####superfluous
            to_full_list.to_csv(f'{metadata_folder}Newly Added to Full List.csv', quotechar='"')
            #SORT COLUMN
            full_list.loc[:, 'sort'] = full_list['Site Name'].apply(lambda x: x.lower().replace('the ', '') if x[:3].lower() == 'the' else x.lower())
            full_list.loc[:, 'sort'] = full_list['sort'].apply(lambda x: x.strip(string.punctuation))
            full_list = full_list.sort_values('sort')
            full_list = full_list.reset_index(drop=True)
            break
        except Exception as e:
            logger.exception('Full List.xlsx update failed.')
            input('Make sure all files are closed, hit enter when they are:>')


    full_list.to_excel('Full List.xlsx', index=False, header=True)


### Commit Changes
    logger.debug('user Decision- Commit Changes?')
    while 'commit' not in input('''
Updates configured to Full List.xlsx and not_active.csv. 
You can check these updates before committing.

        Type "commit" to save and push update. Hit enter to UNDO process.>''').lower():
        if input(f'''\nWARNING: Edited Site Names and Categories will be lost 
WARNING: unless '{metadata_folder}Verified New sites.xlsx' is saved elsewhere.
        
        Type "confirm" to UNDO whole process.>''').lower() == 'confirm':
            raise Exception('Reverting to Prior State')

    logger.debug('Pushing changes to Git Master.')
    os.system('git add "Full List.xlsx" not_active.csv')
    os.system(f'git commit -m "Updated for {folder[16:-1]}"')
    os.system('git push')

#### Generate HTML
    logger.debug('User decision- create HTML?')
    if input('\n   Generate HTML?>[y/n]').lower() == 'y':
        try:
            logger.debug(f'Creating HTML at {folder}A-Z-HTML.html')
            os.system(f'python generateHTML.py "{folder}A-Z-HTML.html"')
            print(f'HTML located in "{folder}A-Z-HTML.html"')
        except:
            logger.exception('''Failed to Produce HTML.
Follow Instructions on WIKI to generate HTML outside of this process.
(https://ukgwa.atlassian.net/wiki/spaces/GUID/pages/1169620993/Creating+A-Z+list+HTML)''')

    print('PROCESS COMPLETE')

### Reverts to prior state
except Exception as e:
    logger.exception('Process Errored or intentionally Interrupted. Rolling Back')
    logger.info('Rolling back To original state.')
    copy(f'{metadata_folder}Full List prev.xlsx', 'Full List.xlsx')
    copy(f'{metadata_folder}Harvesting Summary.csv', source)
    copy(f'{metadata_folder}not_active prev.csv', 'not_active.csv')
    [x.close() for x in logger.handlers]
    rmtree(folder)
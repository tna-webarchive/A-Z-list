import os, sys, re
import pandas as pd
import requests
from tqdm import tqdm
import openpyxl as pxl
from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copy, rmtree
from openpyxl.worksheet.datavalidation import DataValidation
import urllib.parse
from bs4 import BeautifulSoup

folder = 'New URLs between ' +  input('New URLs between>') + '/'
metadata_folder = folder+'metadata/'

#SET UP
while not os.path.isfile('Harvesting Summary.csv'):
    print('Unable to locate Harvesting Summary.csv file.')
    input(f'Please add Harvesting Summary.csv to folder {folder} and hit enter:>')

if not os.path.isdir(folder):
    os.mkdir(folder)

if not os.path.isdir(metadata_folder):
    os.mkdir(metadata_folder)

os.replace('Harvesting Summary.csv', f'{metadata_folder}Harvesting Summary.csv')

def UKGWA_URL(url):
    social = {'twitter.com': 'twitter',
              'flickr.com': 'flickr',
              'youtube.com': 'video'}

    domain = re.search('.*:\/\/(?:www.)?([^\/]+)', url).group(1)

    if domain in social:
        channel = url[:-1].split('/')[-1] + url[-1]     ### takes last character off before splitting in case it is a slash '/'. Appends after split
        return f'https://webarchive.nationalarchives.gov.uk/{social[domain]}/{channel}'
    else:
        return f'https://webarchive.nationalarchives.gov.uk/*/{url}'


def date_range(row):
    if '/*/' in row['Archive URL']:
        query = f'https://webarchive.nationalarchives.gov.uk/largefiles-cdx?url={urllib.parse.quote(row["URL"])}&fields=timestamp&limit=1'
        r = requests.get(query)
        if r.status_code == 200:
            year = r.text[:4]
            range = year + ' - ongoing'
            return range
    else:
        try:
            r = requests.get(row['Archive URL'])
            soup = BeautifulSoup(r.text, 'html.parser')
            tag = soup.find(attrs={'id': 'year_from'})
            tag = tag.find('option')
            year = tag.text
            range = year + ' - ongoing'
            return range
        except:
            pass
    return False

####Load and clean Harvesting Summary
new_sites = pd.read_csv(f'{metadata_folder}Harvesting Summary.csv')
new_sites['Additional Information'] = ''
new_sites.rename(columns={'textbox20': 'URL', 'textbox22': 'Site Name', 'textbox26': 'Archivist Notes', 'Dept_acronym': 'Department'}, inplace=True)
new_sites['Archive URL'] = new_sites['URL'].apply(UKGWA_URL)


####CHECK URLs
frames = [new_sites] + [pd.read_csv(file) for file in os.listdir() if 'not_active' in file]
to_check = pd.concat(frames, ignore_index=True)
to_check.drop_duplicates(subset='URL', inplace=True)
tqdm.pandas(desc='Checking URLs against Archive')       #Creates tqdm instance for progress bar
to_check['Date Range'] = to_check.progress_apply(date_range, axis=1)
to_check.to_csv(f'{metadata_folder}checked.csv', index=False)

#####CREATE NOT ACTIVE CSV
unactive = to_check[to_check['Date Range'] == False].drop(['Date Range'], axis=1)
unactive.to_csv('not_active.csv', index=False)

######CREATE ACTIVE URLs XLSX
active = to_check[to_check['Date Range'] != False]
active = active[['Archive URL', 'Site Name', 'Date Range', 'Additional Information', 'Archivist Notes', 'Department']]  ####define columns
wb = pxl.Workbook()
ws = wb.active
for r in dataframe_to_rows(active, index=False, header=True):
    ws.append(r)

######VERIFY SITE NAMES
wb.save(f'{folder}Verify Site Names.xlsx')
input(f'Verify site names in "{folder}Verify Site Names.xlsx".\nWhen finished, save and close the spreadsheet and hit enter here:>')
input('Re-hit enter to confirm:>')
os.rename(f'{folder}Verify Site Names.xlsx', f'{folder}Cataloguing.xlsx')


#####CREATE LIST WITH EMPTY CATEGORIES
wb = pxl.Workbook()
ws = wb.active
active = pd.read_excel(f'{folder}Cataloguing.xlsx')
for x in range(1, 7):
    active[f'Category #{x}'] = ''
for r in dataframe_to_rows(active, index=False, header=True):
    ws.append(r)

#### CREATE DATA VALIDATION FOR CATEGORIES
cats = wb.create_sheet('cats')
cats.sheet_state = 'hidden'
categories = ['Business, industry, economics and finance', 'Central and regional government', 'Culture and leisure', 'Environment', 'Health, well-being and care', 'Home affairs, public order, justice and rights', 'Honours, awards and appointments', 'International affairs and defence', 'People, community and housing', 'Public inquiries, inquests and royal commissions', 'Transport, communication, science and technology', 'Work, education and skills']
for i, x in enumerate(categories):
    cats.cell(row=i+1, column=1).value = x
dv = DataValidation(type="list", formula1=f'cats!$A$1:$A${i+1}', allow_blank=True)


ws.add_data_validation(dv)
dv.add('G2:L1048576')
wb.save(f'{folder}Add Categories.xlsx')

input(f'Add categories to sites in {folder}Add Categories.xlsx\nWhen finished, close and save spreadsheet and hit enter here:>')
input('Re-hit enter to confirm:>')

###ADD NEW LIST TO FULL LIST
copy('Full List.xlsx', f'{metadata_folder}Full List before new entries.xlsx')
active = pd.read_excel(f'{folder}Add Categories.xlsx')
#active.drop(['Additional Information', 'Archivist Notes'], axis=1, inplace = True)
frames = [pd.read_excel(f'Full List.xlsx')] + [active]
full_list = pd.concat(frames, ignore_index=True)
full_list.drop_duplicates(subset='Archive URL', keep='last')

os.replace(f'{folder}Add Categories.xlsx', f'{metadata_folder}Newly Added to Full List.xlsx')
#SORT COLUMN
full_list['sort'] = full_list['Site Name'].apply(lambda x: x.lower().replace('the ', '') if x[:3].lower() == 'the' else x.lower())
full_list = full_list.sort_values('sort')
full_list = full_list.reset_index(drop=True)

#####Wrtie new full list
wb = pxl.Workbook()
ws = wb.active
for r in dataframe_to_rows(full_list, index=False, header=True):
    ws.append(r)

wb.save(f'Full List.xlsx')


#####WRITE HTML
ABC = 'abcdefghijklmnopqrstuvwxyz'
head = '<h2 name="{lower}">{upper}</h2>'
body = ' <li><a title="This link opens in a new window" href="{URL}" target="_blank" rel="noopener noreferrer">{name}</a></li>'

with open('HTML_template.txt', 'r') as source:
    text = source.read()

for letter in ABC:
    section = full_list[full_list['sort'].str.startswith(letter)]
    if len(section) > 0:
        text += head.replace('{lower}', letter).replace('{upper}', letter.upper())
        text += '\n<ul>\n\n'
        lines = section.apply(lambda x: body.replace('{URL}', x['Archive URL']).replace('{name}', x['Site Name']), axis=1)
        text +='\n'.join(lines)
        text +='\n\n</ul>\n'
    full_list = full_list[~full_list['sort'].str.startswith(letter)]

### 0-9 section
text += head.replace('{lower}', '0-9').replace('{upper}', '0-9')
text += '\n<ul>\n\n'
lines = full_list.apply(lambda x: body.replace('{URL}', x['Archive URL']).replace('{name}', x['Site Name']), axis=1)
text += '\n'.join(lines)
text += '\n\n</ul>'


with open(f'{folder}A-Z list HTML.txt', 'w', encoding='utf-8') as dest:
    dest.write(text)


final = input('You can now check the results. If you\'d like to undo the process, type \'undo\'')

if final.lower() == 'undo':
    os.remove('Full List.xlsx')
    copy(metadata_folder + 'Harvesting Summary.csv', 'Harvesting Summary.csv')
    copy(metadata_folder + 'Full List.xlsx', 'Full List.xlsx')
    rmtree(folder)
#clean
else:
    os.system('git add "Full List.xlsx" not_active.csv')
    os.system(f'git commit -m "Updated for {folder[16:]}"')
    os.system('git push')
    print('Process finished')
